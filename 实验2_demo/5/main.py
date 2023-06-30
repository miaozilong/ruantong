import openpyxl
from openpyxl.styles import Alignment,Font,PatternFill
# 获取工作表信息
def get_ws(path):
    # 获取工作簿
    wb = openpyxl.load_workbook(path)
    # 获取活跃的工作表
    ws = wb.active
    max_row = ws.max_row
    max_column = ws.max_column
    print('当前工作表的名称为：',ws.title)
    print('当前工作表的最大行数为：',max_row)
    print('当前工作表的最大列数为：',max_column)
    return ws,wb
# 打印国家信息
def get_country(ws):
    country_list = []
    # 获取第一列数据，获取都有哪些国家
    cell_col = ws['A']
    print('打印国家名单：')
    for i in cell_col:
        print(i.value,end='\t')
        country_list.append(i.value)
    print('\n----------------分割线----------------')
    return country_list
# 获取某个国家的GDP数据
def get_data(ws,country_name):
    list1 = []
    list2 = []
    # 获取第一行数据
    cell_row = ws['1']
    # 获取列的标题存到list1中
    for i in cell_row:
        list1.append(i.value)
    # 获取第一列数据，获取都有哪些国家
    cell_col = ws['A']
    for index,item in enumerate(cell_col):
        if item.value == country_name:
            cell_row = ws[f'{index+1}']
            # 获取国家从2000到2017年的GDP值
            for i in cell_row:
                list2.append(i.value)
    data = list(zip(list1,list2))
    return data
# 保存数据
def save(wb,data,path,country_name):
    country = wb.create_sheet(country_name)
    for i in range(len(data)):
        for j in range(len(data[0])):
            country.cell(row=i+1,column=j+1).value = data[i][j]
            # 设置首行行高
            country.row_dimensions[1].height = 50
            # 设置列宽
            country.column_dimensions['A'].width = 40
            country.column_dimensions['B'].width = 40
    # 设置格式
    for col in country.columns:
        for cell in col:
            # 设置对齐方式
            cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            # 设置字体
            cell.font = Font(size=10, name='微软雅黑')
    country['A1'].fill = PatternFill(patternType='solid', start_color='FFE6FF')
    country['B1'].fill = PatternFill(patternType='solid', start_color='FFE6FF')
    country['B1'].font = Font(size=15,bold=True, name='微软雅黑')
    wb.save(path)
# 程序入口
def start():
    # 待读取的excel地址
    read_path = 'GDP.xlsx'
    # 处理后的文件保存地址
    save_path = 'E:\\gdp3.xlsx'
    try:
        # 读取文件，获取工作表
        ws,wb = get_ws(read_path)
        # 获取所有国家信息
        country_list = get_country(ws)
        # 获取指定国家的GDP数据
        country_name = input('请输入国家：')
        if country_name not in country_list:
            print('注意：您输入的国家有误！！！')
        else:
            data = get_data(ws,country_name)
            # 保存数据
            save(wb,data,save_path,country_name)
            print('新表格已保存至：',save_path)
    except FileNotFoundError as e:
        print(e)
        print('您打开的文件不存在')
# 调用入口程序
if __name__ == '__main__':
    start()
